from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import CallbackQueryHandler, ConversationHandler, MessageHandler, filters
import openpyxl
import os
import calendar
from datetime import datetime, timedelta
import json

sheet_path = 'sheet.xlsx'
user_data = 'user_data.json'
workbook_cache = None
ADD_SHEET_NAME = range(1)

def create_xlsx_file(file_name: str, title: str = "User") -> None:
    """Create a new xlsx file with the given title and file name."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet['A1'] = 'ID'
    sheet['A2'] = 'Username'
    sheet['A3'] = 'Type of Data'
    sheet['A4'] = 'Message'
    workbook.save(file_name)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """"Start" command handler."""
    user_id = update.effective_user.id
    chat_member = await context.bot.get_chat_member(chat_id='-4754069765', user_id=user_id)

    if chat_member.status in ['member', 'administrator', 'creator']:
        keyboard = [
            [InlineKeyboardButton("Sheets list", callback_data='1')],
            [InlineKeyboardButton("Add sheet", callback_data='2')],
            [InlineKeyboardButton("Delete sheet", callback_data='3')],
            [InlineKeyboardButton("Delete sheets", callback_data='4')],
            [InlineKeyboardButton("Show users", callback_data='5')],
            [InlineKeyboardButton("Show tableusers", callback_data='6')],
            [InlineKeyboardButton("Add user", callback_data='8')],
            [InlineKeyboardButton("Delete user", callback_data='9')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        if not os.path.exists(sheet_path):
            create_xlsx_file(sheet_path)
        if update.message:
            await update.message.reply_text('Hello admin\nwhat you\'d like to do?', reply_markup=reply_markup)
        elif update.callback_query:
            await update.callback_query.edit_message_text('Hello admin\nwhat you\'d like to do?', reply_markup=reply_markup)
        
    else:
        if update.message:
            await update.message.reply_text('You are not a member of the group.')
        elif update.callback_query:
            await update.callback_query.edit_message_text('You are not a member of the group.')

def back_button_markup():
    keyboard = [[InlineKeyboardButton("Back", callback_data='back')]]
    return InlineKeyboardMarkup(keyboard)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(text="Operation cancelled.", reply_markup=back_button_markup())
    return ConversationHandler.END

async def sheets_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global workbook_cache
    if workbook_cache is None:
        workbook_cache = openpyxl.load_workbook(sheet_path)
    workbook = workbook_cache
    sheet_names = [sheet for sheet in workbook.sheetnames if sheet != 'User']
    sheets_text = "\n".join(sheet_names) if sheet_names else "No sheets available."
    await update.callback_query.edit_message_text(text=f"Sheets list:\n{sheets_text}", reply_markup=back_button_markup())
    #await update.callback_query.edit_message_text(text="Sheets list selected", reply_markup=back_button_markup())
    
async def add_sheet(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if 'awaiting_sheet_name' in context.user_data and context.user_data['awaiting_sheet_name']:
        context.user_data['sheet_name'] = update.message.text
        keyboard = [
            [InlineKeyboardButton("One-time", callback_data='one_time')],
            [InlineKeyboardButton("Full-time", callback_data='full_time')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(text="Will this sheet be one-time or full-time?", reply_markup=reply_markup)
        context.user_data['awaiting_sheet_name'] = False
        context.user_data['awaiting_sheet_type'] = True
    elif 'awaiting_sheet_type' in context.user_data and context.user_data['awaiting_sheet_type']:
        query = update.callback_query
        context.user_data['sheet_type'] = query.data
        await query.answer()
        await query.edit_message_text(text="Please enter the IDs of the users to include in the sheet (space-separated):")
        context.user_data['awaiting_sheet_type'] = False
        context.user_data['awaiting_user_ids'] = True
    elif 'awaiting_user_ids' in context.user_data and context.user_data['awaiting_user_ids']:
        user_ids = [int(id.strip()) for id in update.message.text.split()]
        sheet_name = context.user_data['sheet_name']
        if context.user_data['sheet_type'] == 'full_time':
            sheet_name += f" ({datetime.today().strftime('%d.%m.%Y')})"
        workbook = openpyxl.load_workbook(sheet_path)
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sheet_name)
            # Populate the sheet with user information
            for idx, user_id in enumerate(user_ids, start=3):
                sheet.cell(row=1, column=idx, value=user_id)
                user_name = get_user_name_by_id(user_id)  # Implement this function to get the user name by ID
                sheet.cell(row=2, column=idx, value=user_name)
            # Populate the sheet with dates and weekdays
            today = datetime.today()
            first_day = today.replace(day=1)
            last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
            current_day = first_day
            row = 3
            while current_day <= last_day:
                sheet.cell(row=row, column=1, value=current_day.strftime('%d.%m.%Y'))
                sheet.cell(row=row, column=2, value=current_day.strftime('%A'))
                current_day += timedelta(days=1)
                row += 1
            workbook.save(sheet_path)
            await update.message.reply_text(text=f"Sheet '{sheet_name}' created successfully.", reply_markup=back_button_markup())
        else:
            await update.message.reply_text(text=f"Sheet '{sheet_name}' already exists.", reply_markup=back_button_markup())
        context.user_data['awaiting_user_ids'] = False
    else:
        if update.callback_query:
            query = update.callback_query
            await query.answer()
            await query.edit_message_text(text="Please enter a name for the new sheet:")
        else:
            await update.message.reply_text(text="Please enter a name for the new sheet:")
        context.user_data['awaiting_sheet_name'] = True

def get_user_name_by_id(user_id):
    workbook = openpyxl.load_workbook(sheet_path)
    sheet = workbook['User']
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == user_id:
            return sheet.cell(row=2, column=col).value
    return None

async def delete_sheet(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if 'awaiting_sheet_name' in context.user_data and context.user_data['awaiting_sheet_name']:
        sheet_name = update.message.text
        workbook = openpyxl.load_workbook(sheet_path)
        if sheet_name in workbook.sheetnames and sheet_name != 'User':
            workbook.remove(workbook[sheet_name])
            workbook.save(sheet_path)
            await update.message.reply_text(text=f"Sheet '{sheet_name}' deleted successfully.", reply_markup=back_button_markup())
        else:
            await update.message.reply_text(text=f"Sheet '{sheet_name}' cannot be deleted or does not exist.", reply_markup=back_button_markup())
        context.user_data['awaiting_sheet_name'] = False
        context.user_data['delete_sheet'] = False
    else:
        workbook = openpyxl.load_workbook(sheet_path)
        sheet_names = [sheet for sheet in workbook.sheetnames if sheet != 'User']
        sheets_text = "\n".join(sheet_names) if sheet_names else "No sheets available."
        await update.callback_query.edit_message_text(text=f"Sheets list:\n{sheets_text}\nPlease enter the name of the sheet to delete:")
        context.user_data['awaiting_sheet_name'] = True
        context.user_data['delete_sheet'] = True

async def delete_sheets(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.callback_query.edit_message_text(text="Delete sheets selected", reply_markup=back_button_markup())

async def show_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    workbook = openpyxl.load_workbook(sheet_path)
    sheet = workbook['User']
    users_text = ""
    for col in range(1, sheet.max_column + 1):
        user_id = sheet.cell(row=1, column=col).value
        user_name = sheet.cell(row=2, column=col).value
        users_text += f"ID: {user_id}, Name: {user_name}\n"
    await update.callback_query.edit_message_text(text=f"Users list:\n{users_text}", reply_markup=back_button_markup())

async def show_tableusers(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if 'awaiting_table_selection' in context.user_data and context.user_data['awaiting_table_selection']:
        table_index = int(update.message.text) - 1
        workbook = openpyxl.load_workbook(sheet_path)
        sheet_names = [sheet for sheet in workbook.sheetnames if sheet != 'User']
        if 0 <= table_index < len(sheet_names):
            selected_sheet = sheet_names[table_index]
            sheet = workbook[selected_sheet]
            users_text = ""
            for col in range(3, sheet.max_column + 1):
                user_id = sheet.cell(row=1, column=col).value
                user_name = sheet.cell(row=2, column=col).value
                users_text += f"ID: {user_id}, Name: {user_name}\n"
            await update.message.reply_text(text=f"Users in sheet '{selected_sheet}':\n{users_text}", reply_markup=back_button_markup())
        else:
            await update.message.reply_text(text="Invalid selection. Please try again.", reply_markup=back_button_markup())
        context.user_data['awaiting_table_selection'] = False
    else:
        workbook = openpyxl.load_workbook(sheet_path)
        sheet_names = [sheet for sheet in workbook.sheetnames if sheet != 'User']
        sheets_text = "\n".join([f"{i+1}. {sheet}" for i, sheet in enumerate(sheet_names)])
        await update.callback_query.edit_message_text(text=f"Tables list:\n{sheets_text}\nPlease enter the number of the table to view users:")
        context.user_data['awaiting_table_selection'] = True

def load_user_data():
    if not os.path.exists(user_data):
        with open(user_data, 'w') as f:
            json.dump({}, f)
    with open(user_data, 'r') as f:
        return json.load(f)

def save_user_data(data):
    with open(user_data, 'w') as f:
        json.dump(data, f, indent=4)

async def add_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if 'awaiting_user_name' in context.user_data and context.user_data['awaiting_user_name']:
        user_name = update.message.text
        telegram_name = update.effective_user.username
        data = load_user_data()
        if user_name not in data:
            data[user_name] = {"telegram_name": telegram_name, "sheets": []}
            save_user_data(data)
            # Write user data to the Excel sheet
            workbook = openpyxl.load_workbook(sheet_path)
            sheet = workbook['User']
            next_col = sheet.max_column + 1
            sheet.cell(row=1, column=next_col, value=next_col - 1)  # Assuming ID is incremental
            sheet.cell(row=2, column=next_col, value=user_name)
            sheet.cell(row=3, column=next_col, value='User Data')  # Placeholder for type of data
            sheet.cell(row=4, column=next_col, value='Message')  # Placeholder for message
            workbook.save(sheet_path)
            await update.message.reply_text(text=f"User '{user_name}' added successfully.", reply_markup=back_button_markup())
        else:
            await update.message.reply_text(text=f"User '{user_name}' already exists.", reply_markup=back_button_markup())
        context.user_data['awaiting_user_name'] = False
    else:
        if update.callback_query:
            query = update.callback_query
            await query.answer()
            await query.edit_message_text(text="Please enter the name of the user:")
        else:
            await update.message.reply_text(text="Please enter the name of the user:")
        context.user_data['awaiting_user_name'] = True

async def delete_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if 'awaiting_user_id' in context.user_data and context.user_data['awaiting_user_id']:
        user_id = int(update.message.text)
        workbook = openpyxl.load_workbook(sheet_path)
        sheet = workbook['User']
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == user_id:
                sheet.delete_cols(col)
                workbook.save(sheet_path)
                await update.message.reply_text(text=f"User with ID '{user_id}' deleted successfully.", reply_markup=back_button_markup())
                context.user_data['awaiting_user_id'] = False
                return
        await update.message.reply_text(text=f"No user found with ID '{user_id}'.", reply_markup=back_button_markup())
        context.user_data['awaiting_user_id'] = False
    else:
        workbook = openpyxl.load_workbook(sheet_path)
        sheet = workbook['User']
        users_text = ""
        for col in range(1, sheet.max_column + 1):
            user_id = sheet.cell(row=1, column=col).value
            user_name = sheet.cell(row=2, column=col).value
            users_text += f"ID: {user_id}, Name: {user_name}\n"
        await update.callback_query.edit_message_text(text=f"Users list:\n{users_text}\nPlease enter the ID of the user to delete:")
        context.user_data['awaiting_user_id'] = True

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    
    query_data_to_function = {
        '1': sheets_list,
        '2': add_sheet,
        '3': delete_sheet,
        '4': delete_sheets,
        '5': show_users,
        '6': show_tableusers,
        '7': show_tableusers,
        '8': add_user,
        '9': delete_user,
        'back': start,
        'one_time': add_sheet,
        'full_time': add_sheet,
    }

    if query.data in query_data_to_function:
        await query_data_to_function[query.data](update, context)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if 'awaiting_sheet_name' in context.user_data and context.user_data['awaiting_sheet_name']:
        if context.user_data.get('delete_sheet', False):
            await delete_sheet(update, context)
        else:
            await add_sheet(update, context)
    elif 'awaiting_sheet_type' in context.user_data and context.user_data['awaiting_sheet_type']:
        await add_sheet(update, context)
    elif 'awaiting_user_ids' in context.user_data and context.user_data['awaiting_user_ids']:
        await add_sheet(update, context)
    elif 'awaiting_user_name' in context.user_data and context.user_data['awaiting_user_name']:
        await add_user(update, context)
    elif 'awaiting_user_id' in context.user_data and context.user_data['awaiting_user_id']:
        await delete_user(update, context)
    elif 'awaiting_table_selection' in context.user_data and context.user_data['awaiting_table_selection']:
        await show_tableusers(update, context)

app = ApplicationBuilder().token("7717253029:AAGxFbbHnOS-Bct6p9lVK9BgnY8on0opbjo").build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CallbackQueryHandler(button))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

app.run_polling()
